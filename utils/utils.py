import openpyxl as op
import win32com.client

from pathlib import Path

from dateutil import rrule

from constants import file_locations as fl
from constants import dates, holidays
from constants import excel_file as ef


def get_class_names(source_file=fl.CLASS_LIST):
    workbook = op.load_workbook(source_file)
    return workbook.sheetnames


def add_date_to_a_work_sheet(sheet, date):
    sheet.title = date.strftime(ef.DATE_FORMAT)
    sheet[ef.CELL_FOR_DATE] = sheet.title


def add_list_of_learners_to_a_work_sheet(destination_sheet, source_sheet):
    row_start = ef.CLASS_LIST_STARTS_AT_ROW
    row_end = ef.CLASS_LIST_ENDS_AT_ROW
    for row in range(row_start, row_end + 1):
        if source_sheet[f"{ef.NUMBER_COLUMN}{row - row_start + 1}"].value:
            destination_sheet[f"{ef.NUMBER_COLUMN}{row}"].value = \
                source_sheet[f"{ef.NUMBER_COLUMN}{row - row_start + 1}"].value

            destination_sheet[f"{ef.NAME_COLUMN}{row}"].value = \
                source_sheet[f"{ef.SURNAME_COLUMN}{row - row_start + 1}"].value \
                + " " \
                + source_sheet[f"{ef.NAME_COLUMN}{row - row_start + 1}"].value


def set_new_work_sheet_for_each_day(template, class_list_sheet):
    template_sheet = template.active

    for day in rrule.rrule(freq=rrule.DAILY,
                           dtstart=dates.START_DATE,
                           until=dates.END_DATE,
                           byweekday=[rrule.MO, rrule.TU, rrule.WE, rrule.TH, rrule.FR]):
        if day not in holidays.HOLIDAYS_AS_DATE:
            sheet = template.copy_worksheet(template_sheet)
            add_date_to_a_work_sheet(sheet, day)
            add_list_of_learners_to_a_work_sheet(destination_sheet=sheet, source_sheet=class_list_sheet)


def create_class_register_as_excel(class_name, class_list_file):
    template = op.load_workbook(fl.EXCEL_TEMPLATE)

    class_list_wb = op.load_workbook(class_list_file)
    class_list_sheet = class_list_wb[class_name]

    set_new_work_sheet_for_each_day(template, class_list_sheet)
    # for day in rrule.rrule(freq=rrule.DAILY,
    #                        dtstart=dates.START_DATE,
    #                        until=dates.END_DATE,
    #                        byweekday=[rrule.MO, rrule.TU, rrule.WE, rrule.TH, rrule.FR]):
    #     if day not in holidays.HOLIDAYS_AS_DATE:
    #         sheet = template.copy_worksheet(template_sheet)
    #         add_date_to_a_work_sheet(sheet, day)
    #
    #         add_list_of_learners_to_a_work_sheet(destination_sheet=sheet, source_sheet=class_list_sheet)

    template.save(fl.OUTPUT_FOLDER_FOR_EXCEL_FILES / f"{class_name}.xlsx")


def get_pdf_file_name_from_excel_file_name(file_name, destination_folder):
    file_stem = Path(file_name).stem
    pdf_file = destination_folder / f"{file_stem}.pdf"
    return str(pdf_file)


def get_workbook(application, file_name):
    return application.Workbooks.Open(file_name)


def adjust_print_area(workbook):
    for work_sheet in workbook.Sheets:
        work_sheet.PageSetup.Zoom = False
        work_sheet.PageSetup.FitToPagesTall = 1
        work_sheet.PageSetup.FitToPagesWide = 1
        work_sheet.PageSetup.PrintArea = ef.PRINT_AREA


def select_worksheets_to_convert(workbook):
    last_worksheet_index = workbook.Sheets.count
    work_sheet_index_list = list(range(2, last_worksheet_index + 1))
    workbook.WorkSheets(work_sheet_index_list).Select()


def convert_workbook_to_pdf(workbook, file_name, destination_folder=fl.OUTPUT_FOLDER_FOR_PDF_FILES):
    pdf_file_name = get_pdf_file_name_from_excel_file_name(file_name, destination_folder)
    workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_file_name)


def convert_excel_file_to_pdf(file_name, destination_folder=fl.OUTPUT_FOLDER_FOR_PDF_FILES):
    application = win32com.client.Dispatch('Excel.Application')
    application.Visible = False

    workbook = get_workbook(application, file_name)

    adjust_print_area(workbook)

    select_worksheets_to_convert(workbook)

    convert_workbook_to_pdf(workbook, file_name, destination_folder)

    workbook.Close(True)

    application.Quit()


def convert_excel_files_in_a_directory_to_pdf(source_folder=fl.OUTPUT_FOLDER_FOR_EXCEL_FILES,
                                              destination_folder=fl.OUTPUT_FOLDER_FOR_PDF_FILES):
    excel_files = Path(source_folder).resolve().glob('**/*.xlsx')
    for file in excel_files:
        file_name_as_string = str(file)
        convert_excel_file_to_pdf(file_name_as_string, destination_folder)


def initialize_folders(excel_outputs, pdf_outputs):
    Path(excel_outputs).mkdir(parents=True, exist_ok=True)
    Path(pdf_outputs).mkdir(parents=True, exist_ok=True)


def create_class_registers(class_list_file=fl.CLASS_LIST):
    print('Initializing output folders...')
    initialize_folders(excel_outputs=fl.OUTPUT_FOLDER_FOR_EXCEL_FILES, pdf_outputs=fl.OUTPUT_FOLDER_FOR_PDF_FILES)

    print('Getting class names...')
    class_names = get_class_names(class_list_file)

    print('Creating Excel files...')
    for class_name in class_names:
        create_class_register_as_excel(class_name, class_list_file)

    print('Converting Excel files to PDF')
    convert_excel_files_in_a_directory_to_pdf()
